"""Excel (.xlsx) output for scraped jobs."""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import JobDetail

log = logging.getLogger(__name__)

COLUMNS: list[tuple[str, str, int]] = [
    ("Source", "source", 14),
    ("Job ID", "job_id", 16),
    ("Title", "title", 55),
    ("Location", "location", 42),
    ("Country", "country", 22),
    ("Date Posted", "date_posted", 14),
    ("Employment Type", "employment_type", 18),
    ("Hiring Org", "hiring_org", 18),
    ("Detail URL", "detail_url", 70),
    ("Description", "description", 60),
]

_DESCRIPTION_CAP = 32_000  # keep well below Excel's 32,767-char cell limit


def _cell_value(job: JobDetail, attr: str):
    value = getattr(job, attr, None)
    if attr == "description" and isinstance(value, str) and len(value) > _DESCRIPTION_CAP:
        return value[:_DESCRIPTION_CAP] + "\u2026"
    return value


def write_jobs(jobs: Iterable[JobDetail], path: str | Path) -> Path:
    """Write ``jobs`` to an Excel workbook at ``path`` and return the path.

    The workbook has one sheet, ``Jobs``, with a native Excel table for easy
    sorting/filtering. A ``Meta`` sheet records generation time and row count.
    """
    jobs = list(jobs)
    out_path = Path(path).expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = [label for label, _, _ in COLUMNS]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for job in jobs:
        ws.append([_cell_value(job, attr) for _, attr, _ in COLUMNS])

    for col_idx, (_, attr, width) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width

        if attr == "detail_url":
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                url = cell.value
                if isinstance(url, str) and url:
                    cell.hyperlink = url
                    cell.style = "Hyperlink"

        if attr in {"description"}:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

    ws.freeze_panes = "A2"

    if jobs:
        end_col = get_column_letter(len(COLUMNS))
        table_ref = f"A1:{end_col}{ws.max_row}"
        table = Table(displayName="JobsTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    meta = wb.create_sheet("Meta")
    meta["A1"] = "Generated at (UTC)"
    meta["B1"] = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    meta["A2"] = "Job count"
    meta["B2"] = len(jobs)
    meta["A3"] = "Source"
    meta["B3"] = "careers.astrazeneca.com"
    for row in ("1", "2", "3"):
        meta[f"A{row}"].font = Font(bold=True)
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 40

    wb.save(out_path)
    log.info("Wrote %d job(s) to %s", len(jobs), out_path)
    return out_path


__all__ = ["write_jobs", "COLUMNS"]
